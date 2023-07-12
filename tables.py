from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import pandas as pd
import parseTables

alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
            'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

def get_longest_str(d: dict) -> list:
    lengths = []
    for lst in d.values():
        max = 0 
        for text in lst:
            if len(text) > max:
                max = len(text)
        lengths.append(max)
    return lengths


def get_col_max(table: int, index: int, keys: list[list], contents: list[list[list]]) -> int:
    length = len(keys[table][index])
    for row in contents[table]:
        text = row[index]
        length = max(length, len(text))
    return length

def apply_keys_styles(row: int, sheet: object) -> None:
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
        )

    if row == 1:
        font = Font(name = 'Calibri',size = 18, bold = True)
        fill = PatternFill(fill_type = 'solid', start_color = 'E17000')

    else:
        font = Font(name = 'Calibri',size = 16, bold = False)
        fill = PatternFill(fill_type = 'solid', start_color = '0099CCFF')

    for cell in sheet[row]:
        cell.border = border
        cell.font = font
        cell.fill = fill



def get_file(file_name: str) -> None:
    wb = Workbook() 
    keys = parseTables.get_keys(file_name)
    contents = parseTables.get_contents(file_name)

    for i, lst in enumerate(keys):
        temp_sheet = wb.create_sheet(f'Sheet {i + 1}', i)
        temp_sheet.append(lst)
        apply_keys_styles(1, temp_sheet)

        for j, row in enumerate(contents[i], start = 2):
            temp_sheet.append(row)
            apply_keys_styles(j, temp_sheet)

        for k in range(len(keys[i])):
            col = alphabet[k]
            longest = get_col_max(i, k, keys,contents)
            temp_sheet.column_dimensions[col].width = longest + 10

    wb.save('ran.xlsx')


def get_xlsx(file_name: str) -> None:
    data_frames = []
    max_width = []    

    keys = parseTables.get_keys(file_name)
    contents = parseTables.get_contents(file_name)

    #create list of data frames
    for i, table in enumerate(keys):
        dict = {}
        for key in table:
            dict[key] = []

        for row in contents[i]:
            for j, content in enumerate(row):
                key = keys[i][j]
                dict[key].append(content)
        data_frames.append(dict)

    #get string len max to change cell width 
    for d in data_frames:
        widths = get_longest_str(d)
        max_width.append(widths)




    # need to figure out how to get cell wider

    # df = pd.DataFrame(data = data_frames[1])
    # pd.set_option('display.max_columns', None)
    # df.to_excel('kdakdk.xlsx', index=False)


get_file('240.docx')
